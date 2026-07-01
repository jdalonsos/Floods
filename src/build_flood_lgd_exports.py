from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side


DEFAULT_SOURCE_WORKBOOK = Path("data/processed/T20_Anonymised.xlsx")
DEFAULT_JRC_WORKBOOK = Path("data/processed/T20_Anonymised_jrc_flood_check.xlsx")
DEFAULT_GASPAR_WORKBOOK = Path("data/processed/T20_Anonymised_gaspar_check.xlsx")
DEFAULT_HANZE_WORKBOOK = Path("data/processed/T20_Anonymised_hanze_check.xlsx")
DEFAULT_OUTPUT_DIR = Path("outputs/flood_lgd_export")
DEFAULT_SHEET_NAME = "FLOOD_LGD"
DEFAULT_MODE = "copy"
DEFAULT_PROGRESS_EVERY_POINTS = 5_000
DEFAULT_CSV_CHUNK_SIZE = 200_000
DEFAULT_CSV_SEPARATOR = ";"
SOURCE_PRIORITY = ("JRC", "GASPAR", "HANZE")
LATITUDE_ALIASES = ("LAT", "Latitude", "Lat", "Y")
LONGITUDE_ALIASES = ("LONG", "Longitude", "Long", "Lon", "Lng", "X")

OUTPUT_COLUMNS = [
    "point_id",
    "Obligor_ID",
    "Facility_ID",
    "CLOSED_DEFAULT_DATE",
    "Default_Date",
    "ID_ADR",
    "TYPE_ADR",
    "Flag_JRC",
    "Flag_GASPAR",
    "Flag_HANZE",
    "FLOOD_DATA_SOURCE",
    "Flag_JRC_AREA",
    "Flag_GASPAR_AREA",
    "Flag_HANZE_AREA",
    "FLOOD_DATA_SOURCE_AREA",
    "FLAG_FLOOD_ADR",
    "FLAG_FLOOD_ADR_AREA",
    "DATE_REF_FLOOD",
    "DATE_END_FLOOD",
    "FLOOD_DEPTH_MOY",
    "FLOOD_DEPTH_MOY_AREA",
    "FLOOD_DEPTH_MAX",
    "FLOOD_DEPTH_MAX_AREA",
]

COLUMN_WIDTHS = {
    "point_id": 12,
    "Obligor_ID": 16,
    "Facility_ID": 16,
    "CLOSED_DEFAULT_DATE": 22,
    "Default_Date": 18,
    "ID_ADR": 28,
    "TYPE_ADR": 14,
    "Flag_JRC": 11,
    "Flag_GASPAR": 14,
    "Flag_HANZE": 13,
    "FLOOD_DATA_SOURCE": 20,
    "Flag_JRC_AREA": 14,
    "Flag_GASPAR_AREA": 17,
    "Flag_HANZE_AREA": 16,
    "FLOOD_DATA_SOURCE_AREA": 25,
    "FLAG_FLOOD_ADR": 16,
    "FLAG_FLOOD_ADR_AREA": 20,
    "DATE_REF_FLOOD": 18,
    "DATE_END_FLOOD": 18,
    "FLOOD_DEPTH_MOY": 18,
    "FLOOD_DEPTH_MOY_AREA": 20,
    "FLOOD_DEPTH_MAX": 18,
    "FLOOD_DEPTH_MAX_AREA": 20,
}

DATE_FORMAT_COLUMNS = {
    "CLOSED_DEFAULT_DATE",
    "Default_Date",
    "DATE_REF_FLOOD",
    "DATE_END_FLOOD",
}

INTEGER_FORMAT_COLUMNS = {
    "Flag_JRC",
    "Flag_GASPAR",
    "Flag_HANZE",
    "Flag_JRC_AREA",
    "Flag_GASPAR_AREA",
    "Flag_HANZE_AREA",
    "FLAG_FLOOD_ADR",
    "FLAG_FLOOD_ADR_AREA",
}

FLOAT_FORMAT_COLUMNS = {
    "FLOOD_DEPTH_MOY",
    "FLOOD_DEPTH_MOY_AREA",
    "FLOOD_DEPTH_MAX",
    "FLOOD_DEPTH_MAX_AREA",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def log_progress(message: str, *, enabled: bool = True) -> None:
    if not enabled:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {message}", flush=True)


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Build a consolidated FLOOD_LGD export from the checked JRC, Gaspar, and HANZE workbooks. "
            "Rows are emitted at point x consolidated flood-episode level using a 30-day merge rule."
        )
    )
    parser.add_argument("--source-workbook", default=str(DEFAULT_SOURCE_WORKBOOK), help="Original T20 workbook used to recover point metadata such as Obligor_ID and Facility_ID.")
    parser.add_argument("--source-sheet-name", default=None, help="Optional source workbook sheet name. Default uses the first sheet.")
    parser.add_argument("--source-point-id-col", default=None, help="Optional source point identifier column override. Leave blank to auto-detect.")
    parser.add_argument("--source-latitude-col", default=None, help="Optional source latitude column override. Leave blank to auto-detect.")
    parser.add_argument("--source-longitude-col", default=None, help="Optional source longitude column override. Leave blank to auto-detect.")
    parser.add_argument("--source-closed-default-col", default=None, help="Optional source row-end date column override used to fill CLOSED_DEFAULT_DATE.")
    parser.add_argument("--source-closed-default-fallback-col", default=None, help="Optional fallback source row-end date column used when the preferred CLOSED_DEFAULT_DATE column is empty.")
    parser.add_argument("--source-default-date-col", default=None, help="Optional source default-date column override used to fill Default_Date.")
    parser.add_argument("--source-obligor-id-col", default=None, help="Optional source obligor identifier column override.")
    parser.add_argument("--source-facility-id-col", default=None, help="Optional source facility identifier column override.")
    parser.add_argument("--source-type-adr-col", default=None, help="Optional source TYPE_ADR column override.")
    parser.add_argument("--source-type-adr-value", default=None, help="Optional constant value used to fill TYPE_ADR when no source column exists.")
    parser.add_argument("--jrc-workbook", default=str(DEFAULT_JRC_WORKBOOK), help="JRC checked workbook containing candidate_events and event_hits sheets.")
    parser.add_argument("--gaspar-workbook", default=str(DEFAULT_GASPAR_WORKBOOK), help="Gaspar checked workbook containing candidate_events and event_hits sheets.")
    parser.add_argument("--hanze-workbook", default=str(DEFAULT_HANZE_WORKBOOK), help="HANZE checked workbook containing candidate_events and event_hits sheets.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="Folder where the outputs will be written.")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="Name of the derived sheet.")
    parser.add_argument("--merge-gap-days", type=int, default=30, help="Maximum day gap used to merge source rows into one consolidated flood episode. Default: 30.")
    parser.add_argument("--progress-every-points", type=int, default=DEFAULT_PROGRESS_EVERY_POINTS, help=f"Print a clustering progress update every N point_ids/clusters. Default: {DEFAULT_PROGRESS_EVERY_POINTS:,}.")
    parser.add_argument("--csv-chunk-size", type=int, default=DEFAULT_CSV_CHUNK_SIZE, help=f"When --mode csv is used, write the csv in chunks of this many rows so progress can be printed. Default: {DEFAULT_CSV_CHUNK_SIZE:,}.")
    parser.add_argument("--quiet", action="store_true", help="Disable progress logging and only print the final completion message.")
    parser.add_argument(
        "--mode",
        default=DEFAULT_MODE,
        choices=["copy", "standalone", "csv"],
        help=(
            "copy: duplicate the original source workbook and append the new sheet. "
            "standalone: create a small workbook containing only the FLOOD_LGD sheet. "
            "csv: export only the FLOOD_LGD data as a semicolon-separated csv."
        ),
    )
    parser.add_argument("--replace-sheet", action="store_true", help="In copy mode, replace the target sheet if it already exists in the copied workbook.")
    return parser


def parse_args() -> argparse.Namespace:
    return build_argument_parser().parse_args()


def ensure_required_file(path: Path, label: str) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Missing required {label}: {path}")


def normalize_label(value: Any) -> str:
    if pd.isna(value):
        return ""
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum() or ch == "#")


def resolve_column_name(columns: list[str], aliases: tuple[str, ...]) -> str | None:
    normalized_to_original = {
        normalize_label(column): column
        for column in columns
        if normalize_label(column)
    }
    for alias in aliases:
        alias_norm = normalize_label(alias)
        if alias_norm in normalized_to_original:
            return normalized_to_original[alias_norm]
    return None


def build_aliases(requested: str | None, aliases: tuple[str, ...]) -> tuple[str, ...]:
    if requested:
        return (requested, *aliases)
    return aliases


def parse_date(value: Any) -> datetime | None:
    if pd.isna(value) or value in ("", None):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, (int, float, np.integer, np.floating)):
        numeric_value = float(value)
        if np.isnan(numeric_value):
            return None
        parsed = pd.Timestamp("1899-12-30") + pd.to_timedelta(numeric_value, unit="D")
        return parsed.to_pydatetime()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime()
    parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return None
    if isinstance(parsed, pd.Timestamp):
        return parsed.to_pydatetime()
    return parsed


def normalize_point_id_value(value: Any) -> Any:
    if pd.isna(value):
        return pd.NA
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        if np.isnan(value):
            return pd.NA
        if float(value).is_integer():
            return int(value)
        return float(value)
    text = str(value).strip()
    if not text:
        return pd.NA
    try:
        numeric_value = float(text)
    except ValueError:
        return text
    if numeric_value.is_integer():
        return int(numeric_value)
    return text


def normalize_point_id_series(series: pd.Series) -> pd.Series:
    return series.map(normalize_point_id_value)


def normalize_text_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def coerce_bool_series(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        return pd.Series(False, index=df.index)
    series = df[column]
    numeric = pd.to_numeric(series, errors="coerce")
    truthy = numeric.eq(1)
    text = series.astype("string").str.strip().str.lower()
    truthy = truthy | text.isin({"true", "yes", "y", "1"})
    return truthy.fillna(False)


def coordinate_text(value: Any) -> str:
    if pd.isna(value) or value is None:
        return ""
    if isinstance(value, (int, float, np.integer, np.floating)):
        return f"{float(value):.8f}"
    return str(value).strip()


def build_id_adr(latitude_value: Any, longitude_value: Any) -> str | None:
    latitude = coordinate_text(latitude_value)
    longitude = coordinate_text(longitude_value)
    if not latitude and not longitude:
        return None
    return f"{latitude}, {longitude}"


def load_source_frame(
    source_workbook: Path,
    *,
    sheet_name: str | int | None = None,
    point_id_col: str | None = None,
    latitude_col: str | None = None,
    longitude_col: str | None = None,
    closed_default_col: str | None = None,
    closed_default_fallback_col: str | None = None,
    default_date_col: str | None = None,
    obligor_id_col: str | None = None,
    facility_id_col: str | None = None,
    type_adr_col: str | None = None,
    default_type_adr: str | None = None,
) -> pd.DataFrame:
    read_kwargs: dict[str, Any] = {}
    if sheet_name is not None:
        read_kwargs["sheet_name"] = sheet_name
    source_df = pd.read_excel(source_workbook, **read_kwargs)
    if isinstance(source_df, dict):
        if not source_df:
            raise ValueError(f"No sheets found in source workbook: {source_workbook}")
        source_df = next(iter(source_df.values()))
    source_df = source_df.dropna(how="all").copy()
    source_df.columns = [str(column).strip() for column in source_df.columns]

    resolved_point_id_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(point_id_col, ("point_id", "Point ID", "#", "id", "ID_geoloc")),
    )
    if resolved_point_id_col is None:
        source_df["point_id"] = range(1, len(source_df) + 1)
    else:
        source_df["point_id"] = normalize_point_id_series(source_df[resolved_point_id_col])
        missing_mask = source_df["point_id"].isna()
        if missing_mask.any():
            source_df.loc[missing_mask, "point_id"] = range(1, int(missing_mask.sum()) + 1)

    resolved_latitude_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(latitude_col, LATITUDE_ALIASES),
    )
    resolved_longitude_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(longitude_col, LONGITUDE_ALIASES),
    )
    resolved_closed_default_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(
            closed_default_col,
            ("CLOSED_DEFAULT_DATE", "Closed_Default_Date", "Closed Default Date", "last_date", "Last Date"),
        ),
    )
    resolved_closed_default_fallback_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(
            closed_default_fallback_col,
            ("Cut_off_Date", "Cut off Date", "CUT_OFF_DATE"),
        ),
    )
    resolved_default_date_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(default_date_col, ("Default_Date", "Default Date", "DEFAULT_DATE")),
    )
    resolved_obligor_id_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(obligor_id_col, ("Obligor_ID", "Obligor ID")),
    )
    resolved_facility_id_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(facility_id_col, ("Facility_ID", "Facility ID")),
    )
    resolved_type_adr_col = resolve_column_name(
        source_df.columns.tolist(),
        build_aliases(type_adr_col, ("TYPE_ADR", "Type_ADR", "Type ADR")),
    )

    primary_closed_default_dates = (
        source_df[resolved_closed_default_col].map(parse_date)
        if resolved_closed_default_col
        else pd.Series(pd.NaT, index=source_df.index)
    )
    fallback_closed_default_dates = (
        source_df[resolved_closed_default_fallback_col].map(parse_date)
        if resolved_closed_default_fallback_col
        else pd.Series(pd.NaT, index=source_df.index)
    )
    source_df["point_latitude"] = source_df[resolved_latitude_col] if resolved_latitude_col else pd.NA
    source_df["point_longitude"] = source_df[resolved_longitude_col] if resolved_longitude_col else pd.NA
    source_df["CLOSED_DEFAULT_DATE"] = primary_closed_default_dates.combine_first(fallback_closed_default_dates)
    source_df["Default_Date"] = (
        source_df[resolved_default_date_col].map(parse_date)
        if resolved_default_date_col
        else pd.Series(pd.NaT, index=source_df.index)
    )
    source_df["Obligor_ID"] = source_df[resolved_obligor_id_col] if resolved_obligor_id_col else pd.NA
    source_df["Facility_ID"] = source_df[resolved_facility_id_col] if resolved_facility_id_col else pd.NA
    if resolved_type_adr_col:
        source_df["TYPE_ADR"] = source_df[resolved_type_adr_col]
    elif default_type_adr is not None:
        source_df["TYPE_ADR"] = default_type_adr
    else:
        source_df["TYPE_ADR"] = pd.NA
    source_df["ID_ADR"] = source_df.apply(
        lambda row: build_id_adr(row.get("point_latitude"), row.get("point_longitude")),
        axis=1,
    )
    source_df["point_order"] = range(len(source_df))
    return source_df


def read_workbook_sheet(
    workbook_path: Path | None,
    sheet_name: str,
    *,
    label: str,
    verbose: bool,
) -> pd.DataFrame:
    if workbook_path is None or not workbook_path.exists():
        log_progress(f"Skipping missing {label}: {workbook_path}", enabled=verbose)
        return pd.DataFrame()
    log_progress(f"Loading {label} from sheet '{sheet_name}'...", enabled=verbose)
    try:
        df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    except ValueError:
        log_progress(f"Sheet '{sheet_name}' not found in {workbook_path.name}; using an empty {label}.", enabled=verbose)
        return pd.DataFrame()
    result = df.dropna(how="all").copy()
    log_progress(
        f"Loaded {label}: {len(result):,} non-empty rows from {workbook_path.name}.",
        enabled=verbose,
    )
    return result


def standardize_date_bounds(df: pd.DataFrame, start_col: str, end_col: str) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    result = df.copy()
    result[start_col] = result[start_col].map(parse_date) if start_col in result.columns else pd.NaT
    result[end_col] = result[end_col].map(parse_date) if end_col in result.columns else pd.NaT
    if start_col in result.columns and end_col in result.columns:
        result[start_col] = result[start_col].where(result[start_col].notna(), result[end_col])
        result[end_col] = result[end_col].where(result[end_col].notna(), result[start_col])
    return result


def build_jrc_event_frame(event_hits: pd.DataFrame) -> pd.DataFrame:
    if event_hits.empty or "point_id" not in event_hits.columns:
        return pd.DataFrame(columns=_standard_event_columns())
    jrc_df = standardize_date_bounds(event_hits, "start_date", "end_date")
    jrc_df["point_id"] = normalize_point_id_series(jrc_df["point_id"])
    point_flag = coerce_bool_series(jrc_df, "point_buffer_flood_hit").astype(int)
    area_flag = (
        coerce_bool_series(jrc_df, "surrounding_buffer_flood_hit")
        | coerce_bool_series(jrc_df, "buffer_flood_hit")
    ).astype(int)
    jrc_df = jrc_df[point_flag.eq(1) | area_flag.eq(1)].copy()
    if jrc_df.empty:
        return pd.DataFrame(columns=_standard_event_columns())

    jrc_df["source_label"] = "JRC"
    jrc_df["source_priority"] = 0
    jrc_df["source_event_uid"] = jrc_df.get("event_id", pd.Series(range(1, len(jrc_df) + 1), index=jrc_df.index))
    jrc_df["event_start_date"] = jrc_df["start_date"]
    jrc_df["event_end_date"] = jrc_df["end_date"]
    jrc_df["Flag_JRC"] = point_flag.loc[jrc_df.index]
    jrc_df["Flag_GASPAR"] = 0
    jrc_df["Flag_HANZE"] = 0
    jrc_df["Flag_JRC_AREA"] = area_flag.loc[jrc_df.index]
    jrc_df["Flag_GASPAR_AREA"] = 0
    jrc_df["Flag_HANZE_AREA"] = 0
    jrc_df["point_source_active"] = jrc_df["Flag_JRC"].eq(1)
    jrc_df["area_source_active"] = jrc_df["Flag_JRC_AREA"].eq(1)
    jrc_df["jrc_point_depth_mean"] = pd.to_numeric(jrc_df.get("point_buffer_mean_depth_cm"), errors="coerce")
    jrc_df["jrc_area_depth_mean"] = pd.to_numeric(jrc_df.get("buffer_mean_depth_cm"), errors="coerce")
    jrc_df["jrc_point_depth_max"] = pd.to_numeric(jrc_df.get("point_buffer_max_depth_cm"), errors="coerce")
    jrc_df["jrc_area_depth_max"] = pd.to_numeric(jrc_df.get("buffer_max_depth_cm"), errors="coerce")
    return jrc_df[_standard_event_columns()].copy()


def build_fallback_event_frame(
    candidate_df: pd.DataFrame,
    hits_df: pd.DataFrame,
    *,
    source_label: str,
    event_uid_col: str,
    start_col: str,
    end_col: str,
) -> pd.DataFrame:
    if candidate_df.empty or "point_id" not in candidate_df.columns or event_uid_col not in candidate_df.columns:
        return pd.DataFrame(columns=_standard_event_columns())
    fallback_df = candidate_df[candidate_df[event_uid_col].notna()].copy()
    if fallback_df.empty:
        return pd.DataFrame(columns=_standard_event_columns())

    fallback_df = standardize_date_bounds(fallback_df, start_col, end_col)
    fallback_df["point_id"] = normalize_point_id_series(fallback_df["point_id"])
    fallback_df[event_uid_col] = fallback_df[event_uid_col].map(normalize_text_value)

    point_keys: set[tuple[Any, str]] = set()
    if not hits_df.empty and "point_id" in hits_df.columns and event_uid_col in hits_df.columns:
        normalized_hits = hits_df.copy()
        normalized_hits["point_id"] = normalize_point_id_series(normalized_hits["point_id"])
        normalized_hits[event_uid_col] = normalized_hits[event_uid_col].map(normalize_text_value)
        point_keys = {
            (row_point_id, row_event_uid)
            for row_point_id, row_event_uid in zip(
                normalized_hits["point_id"],
                normalized_hits[event_uid_col],
                strict=False,
            )
            if pd.notna(row_point_id) and row_event_uid
        }

    event_keys = list(zip(fallback_df["point_id"], fallback_df[event_uid_col], strict=False))
    point_flag = pd.Series(
        [int((point_id, event_uid) in point_keys) for point_id, event_uid in event_keys],
        index=fallback_df.index,
    )
    area_flag = pd.Series(1, index=fallback_df.index)

    fallback_df["source_label"] = source_label
    fallback_df["source_priority"] = SOURCE_PRIORITY.index(source_label)
    fallback_df["source_event_uid"] = fallback_df[event_uid_col]
    fallback_df["event_start_date"] = fallback_df[start_col]
    fallback_df["event_end_date"] = fallback_df[end_col]
    fallback_df["Flag_JRC"] = 0
    fallback_df["Flag_GASPAR"] = 0
    fallback_df["Flag_HANZE"] = 0
    fallback_df["Flag_JRC_AREA"] = 0
    fallback_df["Flag_GASPAR_AREA"] = 0
    fallback_df["Flag_HANZE_AREA"] = 0
    fallback_df[f"Flag_{source_label}"] = point_flag
    fallback_df[f"Flag_{source_label}_AREA"] = area_flag
    fallback_df["point_source_active"] = point_flag.eq(1)
    fallback_df["area_source_active"] = area_flag.eq(1)
    fallback_df["jrc_point_depth_mean"] = np.nan
    fallback_df["jrc_area_depth_mean"] = np.nan
    fallback_df["jrc_point_depth_max"] = np.nan
    fallback_df["jrc_area_depth_max"] = np.nan
    return fallback_df[_standard_event_columns()].copy()


def _standard_event_columns() -> list[str]:
    return [
        "point_id",
        "source_label",
        "source_priority",
        "source_event_uid",
        "event_start_date",
        "event_end_date",
        "Flag_JRC",
        "Flag_GASPAR",
        "Flag_HANZE",
        "Flag_JRC_AREA",
        "Flag_GASPAR_AREA",
        "Flag_HANZE_AREA",
        "point_source_active",
        "area_source_active",
        "jrc_point_depth_mean",
        "jrc_area_depth_mean",
        "jrc_point_depth_max",
        "jrc_area_depth_max",
    ]


def build_all_event_rows(
    jrc_event_hits: pd.DataFrame,
    gaspar_candidates: pd.DataFrame,
    gaspar_hits: pd.DataFrame,
    hanze_candidates: pd.DataFrame,
    hanze_hits: pd.DataFrame,
    *,
    verbose: bool = False,
) -> pd.DataFrame:
    log_progress("Building source-specific event rows...", enabled=verbose)
    frames = [
        build_jrc_event_frame(jrc_event_hits),
        build_fallback_event_frame(
            gaspar_candidates,
            gaspar_hits,
            source_label="GASPAR",
            event_uid_col="gaspar_event_uid",
            start_col="gaspar_start_date",
            end_col="gaspar_end_date",
        ),
        build_fallback_event_frame(
            hanze_candidates,
            hanze_hits,
            source_label="HANZE",
            event_uid_col="hanze_event_uid",
            start_col="hanze_start_date",
            end_col="hanze_end_date",
        ),
    ]
    non_empty = [frame for frame in frames if not frame.empty]
    if not non_empty:
        log_progress("No positive or candidate flood event rows were built from the source workbooks.", enabled=verbose)
        return pd.DataFrame(columns=_standard_event_columns())
    result = pd.concat(non_empty, ignore_index=True)
    log_progress(
        "Built event rows: "
        f"JRC={len(frames[0]):,}, GASPAR={len(frames[1]):,}, HANZE={len(frames[2]):,}, total={len(result):,}.",
        enabled=verbose,
    )
    return result


def cluster_point_events(point_events: pd.DataFrame, merge_gap_days: int) -> pd.DataFrame:
    if point_events.empty:
        return point_events.copy()
    events = point_events.copy()
    events["event_start_date"] = pd.to_datetime(events["event_start_date"], errors="coerce")
    events["event_end_date"] = pd.to_datetime(events["event_end_date"], errors="coerce")
    events = events.sort_values(
        ["event_start_date", "event_end_date", "source_priority", "source_label", "source_event_uid"],
        kind="stable",
        na_position="last",
    ).reset_index(drop=True)

    cluster_ids: list[int] = []
    current_cluster_id = 1
    current_cluster_end = pd.NaT

    for _, row in events.iterrows():
        start_date = row["event_start_date"]
        end_date = row["event_end_date"]
        if pd.isna(start_date) and pd.isna(end_date):
            if cluster_ids:
                current_cluster_id += 1
            cluster_ids.append(current_cluster_id)
            current_cluster_end = pd.NaT
            continue

        event_start = start_date if pd.notna(start_date) else end_date
        event_end = end_date if pd.notna(end_date) else start_date
        if not cluster_ids:
            cluster_ids.append(current_cluster_id)
            current_cluster_end = event_end
            continue

        if pd.isna(current_cluster_end) or pd.isna(event_start):
            current_cluster_id += 1
            cluster_ids.append(current_cluster_id)
            current_cluster_end = event_end
            continue

        gap_days = (event_start.normalize() - current_cluster_end.normalize()).days
        if gap_days <= merge_gap_days:
            cluster_ids.append(current_cluster_id)
            if pd.notna(event_end) and (pd.isna(current_cluster_end) or event_end > current_cluster_end):
                current_cluster_end = event_end
        else:
            current_cluster_id += 1
            cluster_ids.append(current_cluster_id)
            current_cluster_end = event_end

    events["cluster_id"] = cluster_ids
    return events


def choose_priority_source(row: pd.Series, area: bool = False) -> str | pd.NA:
    suffix = "_AREA" if area else ""
    for label in SOURCE_PRIORITY:
        if int(row.get(f"Flag_{label}{suffix}", 0) or 0) == 1:
            return label
    return pd.NA


def select_source_rows(
    cluster_df: pd.DataFrame,
    source_label: str | pd.NA,
    *,
    point_level: bool,
) -> pd.DataFrame:
    if pd.isna(source_label):
        return cluster_df.iloc[0:0].copy()

    active_column = "point_source_active" if point_level else "area_source_active"
    selected = cluster_df[
        cluster_df["source_label"].eq(source_label)
        & cluster_df[active_column].fillna(False)
    ].copy()
    if selected.empty:
        selected = cluster_df[cluster_df["source_label"].eq(source_label)].copy()
    return selected


def aggregate_cluster_rows(cluster_df: pd.DataFrame) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for label in SOURCE_PRIORITY:
        result[f"Flag_{label}"] = int(cluster_df[f"Flag_{label}"].max()) if f"Flag_{label}" in cluster_df.columns else 0
        result[f"Flag_{label}_AREA"] = int(cluster_df[f"Flag_{label}_AREA"].max()) if f"Flag_{label}_AREA" in cluster_df.columns else 0

    result["FLOOD_DATA_SOURCE"] = choose_priority_source(pd.Series(result), area=False)
    result["FLOOD_DATA_SOURCE_AREA"] = choose_priority_source(pd.Series(result), area=True)
    result["FLAG_FLOOD_ADR"] = int(any(result[f"Flag_{label}"] == 1 for label in SOURCE_PRIORITY))
    result["FLAG_FLOOD_ADR_AREA"] = int(any(result[f"Flag_{label}_AREA"] == 1 for label in SOURCE_PRIORITY))

    point_source = result["FLOOD_DATA_SOURCE"]
    area_source = result["FLOOD_DATA_SOURCE_AREA"]

    result["DATE_REF_FLOOD"] = pd.NaT
    result["DATE_END_FLOOD"] = pd.NaT
    if pd.notna(point_source):
        selected_source_rows = select_source_rows(cluster_df, point_source, point_level=True)
        result["DATE_REF_FLOOD"] = selected_source_rows["event_start_date"].min()
        result["DATE_END_FLOOD"] = selected_source_rows["event_end_date"].max()
    elif pd.notna(area_source):
        selected_area_rows = select_source_rows(cluster_df, area_source, point_level=False)
        result["DATE_REF_FLOOD"] = selected_area_rows["event_start_date"].min()
        result["DATE_END_FLOOD"] = selected_area_rows["event_end_date"].max()

    jrc_point_rows = cluster_df[cluster_df["Flag_JRC"].eq(1)]
    jrc_area_rows = cluster_df[cluster_df["Flag_JRC_AREA"].eq(1)]
    result["FLOOD_DEPTH_MOY"] = (
        pd.to_numeric(jrc_point_rows["jrc_point_depth_mean"], errors="coerce").max()
        if pd.notna(point_source) and point_source == "JRC" and not jrc_point_rows.empty
        else np.nan
    )
    result["FLOOD_DEPTH_MOY_AREA"] = (
        pd.to_numeric(jrc_area_rows["jrc_area_depth_mean"], errors="coerce").max()
        if pd.notna(area_source) and area_source == "JRC" and not jrc_area_rows.empty
        else np.nan
    )
    result["FLOOD_DEPTH_MAX"] = (
        pd.to_numeric(jrc_point_rows["jrc_point_depth_max"], errors="coerce").max()
        if pd.notna(point_source) and point_source == "JRC" and not jrc_point_rows.empty
        else np.nan
    )
    result["FLOOD_DEPTH_MAX_AREA"] = (
        pd.to_numeric(jrc_area_rows["jrc_area_depth_max"], errors="coerce").max()
        if pd.notna(area_source) and area_source == "JRC" and not jrc_area_rows.empty
        else np.nan
    )
    return result


def build_source_export_base(source_df: pd.DataFrame) -> pd.DataFrame:
    base = source_df.copy()
    base["point_id"] = normalize_point_id_series(base["point_id"])
    for column in ["Obligor_ID", "Facility_ID", "CLOSED_DEFAULT_DATE", "Default_Date", "ID_ADR", "TYPE_ADR"]:
        if column not in base.columns:
            base[column] = pd.NA
    if "point_order" not in base.columns:
        base["point_order"] = range(len(base))
    return base[
        [
            "point_id",
            "Obligor_ID",
            "Facility_ID",
            "CLOSED_DEFAULT_DATE",
            "Default_Date",
            "ID_ADR",
            "TYPE_ADR",
            "point_order",
        ]
    ].copy()


def build_flood_lgd_dataframe(
    source_df: pd.DataFrame,
    jrc_event_hits: pd.DataFrame,
    gaspar_candidates: pd.DataFrame,
    gaspar_hits: pd.DataFrame,
    hanze_candidates: pd.DataFrame,
    hanze_hits: pd.DataFrame,
    *,
    merge_gap_days: int = 30,
    progress_every_points: int = DEFAULT_PROGRESS_EVERY_POINTS,
    verbose: bool = False,
) -> pd.DataFrame:
    """Build the final T20 flood export at point x consolidated-flood-episode level.

    Input workbooks stay source-specific:
    - JRC contributes raster-confirmed event hits
    - GASPAR contributes commune candidates plus point-positive hits
    - HANZE contributes department candidates plus point-positive hits

    The final output is one consolidated table. Events across sources are merged
    per `point_id` when their intervals overlap or stay within `merge_gap_days`,
    then source priority `JRC > GASPAR > HANZE` is used for the retained source
    fields. Points with no flood evidence are still emitted once with zero flags
    and missing flood dates.
    """
    base_df = build_source_export_base(source_df)
    log_progress(f"Prepared source base table with {len(base_df):,} point rows.", enabled=verbose)
    event_rows = build_all_event_rows(
        jrc_event_hits=jrc_event_hits,
        gaspar_candidates=gaspar_candidates,
        gaspar_hits=gaspar_hits,
        hanze_candidates=hanze_candidates,
        hanze_hits=hanze_hits,
        verbose=verbose,
    )
    if event_rows.empty:
        log_progress("No flood events remained after source normalization; emitting one zero row per point.", enabled=verbose)
        no_flood_df = base_df.copy()
        for column in OUTPUT_COLUMNS:
            if column not in no_flood_df.columns:
                no_flood_df[column] = pd.NA
        for flag_column in [
            "Flag_JRC",
            "Flag_GASPAR",
            "Flag_HANZE",
            "Flag_JRC_AREA",
            "Flag_GASPAR_AREA",
            "Flag_HANZE_AREA",
            "FLAG_FLOOD_ADR",
            "FLAG_FLOOD_ADR_AREA",
        ]:
            no_flood_df[flag_column] = 0
        no_flood_df["DATE_REF_FLOOD"] = pd.NaT
        no_flood_df["DATE_END_FLOOD"] = pd.NaT
        no_flood_df["FLOOD_DEPTH_MOY"] = np.nan
        no_flood_df["FLOOD_DEPTH_MOY_AREA"] = np.nan
        no_flood_df["FLOOD_DEPTH_MAX"] = np.nan
        no_flood_df["FLOOD_DEPTH_MAX_AREA"] = np.nan
        return no_flood_df[OUTPUT_COLUMNS].copy()

    clustered_frames: list[pd.DataFrame] = []
    point_groups = event_rows.groupby("point_id", sort=False)
    total_points_with_events = point_groups.ngroups
    processed_points = 0
    processed_event_rows = 0
    progress_step = max(1, progress_every_points)
    log_progress(
        f"Clustering {len(event_rows):,} source rows across {total_points_with_events:,} point_ids...",
        enabled=verbose,
    )
    for point_id, point_events in point_groups:
        point_clustered = cluster_point_events(point_events, merge_gap_days=merge_gap_days)
        point_clustered["point_id"] = point_id
        clustered_frames.append(point_clustered)
        processed_points += 1
        processed_event_rows += len(point_events)
        if (
            verbose
            and (
                processed_points == 1
                or processed_points % progress_step == 0
                or processed_points == total_points_with_events
            )
        ):
            log_progress(
                f"Clustered {processed_points:,}/{total_points_with_events:,} point_ids "
                f"({processed_event_rows:,}/{len(event_rows):,} source rows processed).",
                enabled=True,
            )
    clustered_events = pd.concat(clustered_frames, ignore_index=True)
    log_progress(
        f"Finished clustering: {len(clustered_events):,} source rows assigned to clusters.",
        enabled=verbose,
    )

    output_rows: list[dict[str, Any]] = []
    flooded_point_ids = set(clustered_events["point_id"].dropna().tolist())
    metadata_by_point = base_df.set_index("point_id").to_dict(orient="index")

    cluster_groups = clustered_events.groupby(["point_id", "cluster_id"], sort=False)
    total_clusters = cluster_groups.ngroups
    processed_clusters = 0
    log_progress(
        f"Aggregating {total_clusters:,} consolidated flood clusters into final rows...",
        enabled=verbose,
    )
    for (point_id, cluster_id), cluster_df in cluster_groups:
        row = dict(metadata_by_point.get(point_id, {}))
        row["point_id"] = point_id
        row["cluster_id"] = cluster_id
        row.update(aggregate_cluster_rows(cluster_df))
        output_rows.append(row)
        processed_clusters += 1
        if (
            verbose
            and (
                processed_clusters == 1
                or processed_clusters % progress_step == 0
                or processed_clusters == total_clusters
            )
        ):
            log_progress(
                f"Aggregated {processed_clusters:,}/{total_clusters:,} consolidated clusters.",
                enabled=True,
            )

    for point_id, metadata in metadata_by_point.items():
        if point_id in flooded_point_ids:
            continue
        output_rows.append(
            {
                **metadata,
                "point_id": point_id,
                "Flag_JRC": 0,
                "Flag_GASPAR": 0,
                "Flag_HANZE": 0,
                "Flag_JRC_AREA": 0,
                "Flag_GASPAR_AREA": 0,
                "Flag_HANZE_AREA": 0,
                "FLOOD_DATA_SOURCE": pd.NA,
                "FLOOD_DATA_SOURCE_AREA": pd.NA,
                "FLAG_FLOOD_ADR": 0,
                "FLAG_FLOOD_ADR_AREA": 0,
                "DATE_REF_FLOOD": pd.NaT,
                "DATE_END_FLOOD": pd.NaT,
                "FLOOD_DEPTH_MOY": np.nan,
                "FLOOD_DEPTH_MOY_AREA": np.nan,
                "FLOOD_DEPTH_MAX": np.nan,
                "FLOOD_DEPTH_MAX_AREA": np.nan,
            }
        )
    log_progress(
        f"Added no-flood rows for {len(base_df) - len(flooded_point_ids):,} point_ids without any merged flood cluster.",
        enabled=verbose,
    )

    result = pd.DataFrame(output_rows)
    result = result.merge(base_df[["point_id", "point_order"]], on="point_id", how="left", suffixes=("", "_base"))
    result["sort_date"] = pd.to_datetime(result["DATE_REF_FLOOD"], errors="coerce")
    result = result.sort_values(["point_order", "sort_date"], kind="stable", na_position="last").reset_index(drop=True)
    for flag_column in [
        "Flag_JRC",
        "Flag_GASPAR",
        "Flag_HANZE",
        "Flag_JRC_AREA",
        "Flag_GASPAR_AREA",
        "Flag_HANZE_AREA",
        "FLAG_FLOOD_ADR",
        "FLAG_FLOOD_ADR_AREA",
    ]:
        result[flag_column] = pd.to_numeric(result[flag_column], errors="coerce").fillna(0).astype(int)
    log_progress(f"Final FLOOD_LGD dataframe contains {len(result):,} rows.", enabled=verbose)
    return result[OUTPUT_COLUMNS].copy()


def style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    worksheet.row_dimensions[1].height = 22

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER

    header_cells = list(worksheet[1])
    header_to_column_letter = {
        cell.value: cell.column_letter
        for cell in header_cells
        if cell.value
    }
    for column_name, width in COLUMN_WIDTHS.items():
        column_letter = header_to_column_letter.get(column_name)
        if column_letter:
            worksheet.column_dimensions[column_letter].width = width

    max_row = worksheet.max_row
    if max_row < 2:
        return

    for column_name in DATE_FORMAT_COLUMNS:
        column_letter = header_to_column_letter.get(column_name)
        if not column_letter:
            continue
        for column_cells in worksheet[f"{column_letter}2:{column_letter}{max_row}"]:
            for cell in column_cells:
                cell.number_format = "yyyy-mm-dd"

    for column_name in INTEGER_FORMAT_COLUMNS:
        column_letter = header_to_column_letter.get(column_name)
        if not column_letter:
            continue
        for column_cells in worksheet[f"{column_letter}2:{column_letter}{max_row}"]:
            for cell in column_cells:
                cell.number_format = "0"

    for column_name in FLOAT_FORMAT_COLUMNS:
        column_letter = header_to_column_letter.get(column_name)
        if not column_letter:
            continue
        for column_cells in worksheet[f"{column_letter}2:{column_letter}{max_row}"]:
            for cell in column_cells:
                cell.number_format = "0.##"


def normalize_excel_cell(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, np.generic):
        return value.item()
    return value


def write_sheet_into_existing_workbook(
    workbook_path: Path,
    output_path: Path,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    replace_sheet: bool,
) -> None:
    workbook = load_workbook(workbook_path)

    if sheet_name in workbook.sheetnames:
        if not replace_sheet:
            raise ValueError(
                f"Sheet {sheet_name!r} already exists in {workbook_path.name}. "
                "Use --replace-sheet if you want to overwrite it in the copied workbook."
            )
        del workbook[sheet_name]

    worksheet = workbook.create_sheet(title=sheet_name)
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([normalize_excel_cell(value) for value in row])
    style_worksheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_standalone_workbook(
    output_path: Path,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([normalize_excel_cell(value) for value in row])
    style_worksheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csv_output(
    output_path: Path,
    df: pd.DataFrame,
    *,
    chunk_size: int,
    verbose: bool,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    effective_chunk_size = max(1, chunk_size)
    total_rows = len(df)
    log_progress(
        f"Writing semicolon-separated csv output to {output_path} in chunks of {effective_chunk_size:,} rows...",
        enabled=verbose,
    )
    if total_rows == 0:
        df.to_csv(output_path, index=False, encoding="utf-8", sep=DEFAULT_CSV_SEPARATOR)
        log_progress(f"Wrote empty csv with headers to {output_path}.", enabled=verbose)
        return

    first_chunk = True
    for start_idx in range(0, total_rows, effective_chunk_size):
        end_idx = min(start_idx + effective_chunk_size, total_rows)
        df.iloc[start_idx:end_idx].to_csv(
            output_path,
            index=False,
            encoding="utf-8",
            sep=DEFAULT_CSV_SEPARATOR,
            mode="w" if first_chunk else "a",
            header=first_chunk,
        )
        first_chunk = False
        log_progress(
            f"Wrote {end_idx:,}/{total_rows:,} csv rows to {output_path.name}.",
            enabled=verbose,
        )


def build_output_path(source_workbook: Path, sheet_name: str, mode: str) -> Path:
    safe_sheet_name = sheet_name.replace(" ", "_")
    if mode == "copy":
        return Path(f"{source_workbook.stem}_with_{safe_sheet_name}{source_workbook.suffix}")
    if mode == "standalone":
        return Path(f"{source_workbook.stem}_{safe_sheet_name}_only.xlsx")
    if mode == "csv":
        return Path(f"{source_workbook.stem}_{safe_sheet_name}.csv")
    raise ValueError(f"Unsupported mode: {mode}")


def run(args: argparse.Namespace) -> None:
    verbose = not args.quiet

    if not args.source_workbook:
        raise ValueError("Missing required source workbook path.")

    source_workbook = Path(args.source_workbook)
    jrc_workbook = Path(args.jrc_workbook)
    gaspar_workbook = Path(args.gaspar_workbook)
    hanze_workbook = Path(args.hanze_workbook)
    output_dir = Path(args.output_dir)

    ensure_required_file(source_workbook, "source workbook")
    ensure_required_file(jrc_workbook, "JRC workbook")
    ensure_required_file(gaspar_workbook, "Gaspar workbook")
    if not hanze_workbook.exists():
        log_progress(f"HANZE workbook not found at {hanze_workbook}. HANZE columns will stay zero/NA.", enabled=verbose)

    log_progress(f"Loading source workbook from {source_workbook}...", enabled=verbose)
    source_df = load_source_frame(
        source_workbook,
        sheet_name=args.source_sheet_name,
        point_id_col=args.source_point_id_col,
        latitude_col=args.source_latitude_col,
        longitude_col=args.source_longitude_col,
        closed_default_col=args.source_closed_default_col,
        closed_default_fallback_col=args.source_closed_default_fallback_col,
        default_date_col=args.source_default_date_col,
        obligor_id_col=args.source_obligor_id_col,
        facility_id_col=args.source_facility_id_col,
        type_adr_col=args.source_type_adr_col,
        default_type_adr=args.source_type_adr_value,
    )
    log_progress(f"Loaded source workbook with {len(source_df):,} rows.", enabled=verbose)
    jrc_event_hits = read_workbook_sheet(jrc_workbook, "event_hits", label="JRC event hits", verbose=verbose)
    gaspar_candidates = read_workbook_sheet(
        gaspar_workbook,
        "candidate_events",
        label="GASPAR candidate events",
        verbose=verbose,
    )
    gaspar_hits = read_workbook_sheet(gaspar_workbook, "event_hits", label="GASPAR event hits", verbose=verbose)
    hanze_candidates = read_workbook_sheet(
        hanze_workbook,
        "candidate_events",
        label="HANZE candidate events",
        verbose=verbose,
    )
    hanze_hits = read_workbook_sheet(hanze_workbook, "event_hits", label="HANZE event hits", verbose=verbose)

    log_progress("Starting FLOOD_LGD consolidation...", enabled=verbose)
    flood_lgd_df = build_flood_lgd_dataframe(
        source_df=source_df,
        jrc_event_hits=jrc_event_hits,
        gaspar_candidates=gaspar_candidates,
        gaspar_hits=gaspar_hits,
        hanze_candidates=hanze_candidates,
        hanze_hits=hanze_hits,
        merge_gap_days=args.merge_gap_days,
        progress_every_points=args.progress_every_points,
        verbose=verbose,
    )

    output_path = output_dir / build_output_path(source_workbook, args.sheet_name, args.mode)
    if args.mode == "copy":
        log_progress(f"Writing Excel copy output to {output_path}...", enabled=verbose)
        write_sheet_into_existing_workbook(
            source_workbook,
            output_path,
            args.sheet_name,
            flood_lgd_df,
            replace_sheet=args.replace_sheet,
        )
    elif args.mode == "standalone":
        log_progress(f"Writing standalone workbook to {output_path}...", enabled=verbose)
        write_standalone_workbook(output_path, args.sheet_name, flood_lgd_df)
    elif args.mode == "csv":
        write_csv_output(
            output_path,
            flood_lgd_df,
            chunk_size=args.csv_chunk_size,
            verbose=verbose,
        )
    else:
        raise ValueError(f"Unsupported mode: {args.mode}")

    print(
        f"Wrote {output_path} with {len(flood_lgd_df):,} consolidated {args.sheet_name} rows "
        f"using merge_gap_days={args.merge_gap_days}."
    )


def main() -> None:
    args = parse_args()
    run(args)


if __name__ == "__main__":
    main()
