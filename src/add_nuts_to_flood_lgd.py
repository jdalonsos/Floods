from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
import sys
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SITE_PACKAGES = PROJECT_ROOT / ".venv" / "Lib" / "site-packages"
if SITE_PACKAGES.exists() and str(SITE_PACKAGES) not in sys.path:
    sys.path.insert(0, str(SITE_PACKAGES))

import geopandas as gpd
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from build_flood_lgd_exports import (
    DEFAULT_CSV_SEPARATOR,
    DEFAULT_SHEET_NAME,
    ensure_required_file,
    log_progress,
    normalize_excel_cell,
    style_worksheet,
)
from eurostat_nuts_lookup import load_nuts


COORDINATE_PATTERN = re.compile(r"[-+]?\d+(?:\.\d+)?")
NUTS_ENRICHMENT_COLUMNS = [
    "point_latitude",
    "point_longitude",
    "id_adr_coordinate_order",
    "id_adr_order_resolution",
    "nuts1_code",
    "nuts1_name",
    "nuts2_code",
    "nuts2_name",
    "nuts3_code",
    "nuts3_name",
]


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Add NUTS 1/2/3 codes and names to an existing FLOOD_LGD csv/xlsx by "
            "parsing ID_ADR coordinates and spatially joining them to Eurostat NUTS polygons."
        )
    )
    parser.add_argument("--flood-lgd-file", required=True, help="Existing FLOOD_LGD csv or xlsx file to enrich.")
    parser.add_argument(
        "--nuts-file",
        default="NUTS_RG_03M_2024_4326.gpkg",
        help="Official Eurostat NUTS GeoPackage in EPSG:4326. Default: NUTS_RG_03M_2024_4326.gpkg",
    )
    parser.add_argument("--output-file", default=None, help="Optional output path. Default writes a sibling file with a _with_nuts suffix.")
    parser.add_argument("--in-place", action="store_true", help="Overwrite --flood-lgd-file instead of writing a sibling file.")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="Sheet name to update when --flood-lgd-file is an Excel workbook.")
    parser.add_argument("--id-adr-col", default="ID_ADR", help="Column containing the combined coordinate text. Default: ID_ADR.")
    parser.add_argument("--country-code", default="FR", help="Country code used to filter the NUTS dataset. Default: FR.")
    parser.add_argument(
        "--default-coordinate-order",
        choices=["lat_lon", "lon_lat"],
        default="lat_lon",
        help=(
            "Fallback order to use when neither candidate order can be distinguished "
            "from NUTS matches or the country bounding box. Default: lat_lon."
        ),
    )
    parser.add_argument("--quiet", action="store_true", help="Disable progress logging and only print the final completion message.")
    return parser


def derive_output_path(export_path: Path) -> Path:
    return export_path.with_name(f"{export_path.stem}_with_nuts{export_path.suffix}")


def detect_csv_separator(export_path: Path) -> str:
    with export_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return dialect.delimiter
    except csv.Error:
        return DEFAULT_CSV_SEPARATOR


def read_flood_lgd_frame(export_path: Path, *, sheet_name: str, verbose: bool) -> pd.DataFrame:
    suffix = export_path.suffix.lower()
    log_progress(f"Loading existing FLOOD_LGD output from {export_path}...", enabled=verbose)
    if suffix == ".csv":
        separator = detect_csv_separator(export_path)
        log_progress(f"Detected csv separator {separator!r}.", enabled=verbose)
        return pd.read_csv(export_path, sep=separator)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return pd.read_excel(export_path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported FLOOD_LGD file type: {export_path.suffix}")


def parse_id_adr_value(value: Any) -> tuple[float, float]:
    if pd.isna(value) or value is None:
        return (np.nan, np.nan)
    matches = COORDINATE_PATTERN.findall(str(value))
    if len(matches) < 2:
        return (np.nan, np.nan)
    try:
        return (float(matches[0]), float(matches[1]))
    except ValueError:
        return (np.nan, np.nan)


def spatial_join_level(
    points_gdf: gpd.GeoDataFrame,
    level_gdf: gpd.GeoDataFrame,
    *,
    key_col: str,
    code_col: str,
    name_col: str,
) -> pd.DataFrame:
    joined = gpd.sjoin(
        points_gdf[[key_col, "geometry"]],
        level_gdf[[code_col, name_col, "geometry"]],
        how="left",
        predicate="within",
    ).drop(columns=["index_right"], errors="ignore")
    joined = joined.drop_duplicates(subset=[key_col], keep="first")

    missing_mask = joined[code_col].isna()
    if missing_mask.any():
        missing_keys = joined.loc[missing_mask, key_col].tolist()
        fallback = gpd.sjoin(
            points_gdf.loc[points_gdf[key_col].isin(missing_keys), [key_col, "geometry"]],
            level_gdf[[code_col, name_col, "geometry"]],
            how="left",
            predicate="intersects",
        ).drop(columns=["index_right"], errors="ignore")
        fallback = fallback.drop_duplicates(subset=[key_col], keep="first")
        if not fallback.empty:
            fallback = fallback.set_index(key_col)
            joined = joined.set_index(key_col)
            joined.loc[fallback.index, code_col] = fallback[code_col]
            joined.loc[fallback.index, name_col] = fallback[name_col]
            joined = joined.reset_index()

    return joined[[key_col, code_col, name_col]].copy()


def build_candidate_lookup(
    candidate_df: pd.DataFrame,
    *,
    key_col: str,
    latitude_col: str,
    longitude_col: str,
    nuts_gdf: gpd.GeoDataFrame,
) -> pd.DataFrame:
    result = candidate_df[[key_col]].copy()

    valid_mask = candidate_df[latitude_col].notna() & candidate_df[longitude_col].notna()
    if not valid_mask.any():
        for column in ["nuts1_code", "nuts1_name", "nuts2_code", "nuts2_name", "nuts3_code", "nuts3_name"]:
            result[column] = pd.NA
        result["nuts_match_level_count"] = 0
        return result

    points_gdf = gpd.GeoDataFrame(
        candidate_df.loc[valid_mask, [key_col]].copy(),
        geometry=gpd.points_from_xy(
            candidate_df.loc[valid_mask, longitude_col],
            candidate_df.loc[valid_mask, latitude_col],
        ),
        crs="EPSG:4326",
    )

    lookup = points_gdf[[key_col]].copy()
    for level in (1, 2, 3):
        code_col = f"nuts{level}_code"
        name_col = f"nuts{level}_name"
        level_gdf = nuts_gdf.loc[nuts_gdf["LEVL_CODE"] == level, ["NUTS_ID", "NUTS_NAME", "geometry"]].rename(
            columns={"NUTS_ID": code_col, "NUTS_NAME": name_col}
        )
        level_lookup = spatial_join_level(
            points_gdf,
            level_gdf,
            key_col=key_col,
            code_col=code_col,
            name_col=name_col,
        )
        lookup = lookup.merge(level_lookup, on=key_col, how="left")

    lookup["nuts_match_level_count"] = lookup[["nuts1_code", "nuts2_code", "nuts3_code"]].notna().sum(axis=1).astype(int)
    result = result.merge(lookup, on=key_col, how="left")
    result["nuts_match_level_count"] = result["nuts_match_level_count"].fillna(0).astype(int)
    return result


def is_within_country_bounds(
    latitude: pd.Series,
    longitude: pd.Series,
    bounds: tuple[float, float, float, float],
) -> pd.Series:
    minx, miny, maxx, maxy = bounds
    return (
        latitude.notna()
        & longitude.notna()
        & longitude.ge(minx)
        & longitude.le(maxx)
        & latitude.ge(miny)
        & latitude.le(maxy)
    )


def reorder_enrichment_columns(df: pd.DataFrame, *, id_adr_col: str) -> pd.DataFrame:
    present_columns = [column for column in NUTS_ENRICHMENT_COLUMNS if column in df.columns]
    if not present_columns:
        return df

    ordered_columns = [column for column in df.columns if column not in present_columns]
    if id_adr_col in ordered_columns:
        insert_at = ordered_columns.index(id_adr_col) + 1
    elif "TYPE_ADR" in ordered_columns:
        insert_at = ordered_columns.index("TYPE_ADR") + 1
    elif "point_id" in ordered_columns:
        insert_at = ordered_columns.index("point_id") + 1
    else:
        insert_at = len(ordered_columns)
    ordered_columns[insert_at:insert_at] = present_columns
    return df[ordered_columns].copy()


def attach_nuts_metadata(
    export_df: pd.DataFrame,
    nuts_gdf: gpd.GeoDataFrame,
    *,
    id_adr_col: str,
    default_coordinate_order: str,
) -> pd.DataFrame:
    if id_adr_col not in export_df.columns:
        raise KeyError(f"The FLOOD_LGD file must contain an {id_adr_col!r} column.")

    result = export_df.reset_index(drop=True).copy()
    result["__row_id__"] = result.index

    parsed = result[id_adr_col].apply(parse_id_adr_value).tolist()
    parsed_df = pd.DataFrame(parsed, columns=["__coord_1__", "__coord_2__"], index=result.index)
    result = pd.concat([result, parsed_df], axis=1)

    lat_lon_candidate = result[["__row_id__", "__coord_1__", "__coord_2__"]].rename(
        columns={"__coord_1__": "point_latitude", "__coord_2__": "point_longitude"}
    )
    lon_lat_candidate = result[["__row_id__", "__coord_1__", "__coord_2__"]].rename(
        columns={"__coord_1__": "point_longitude", "__coord_2__": "point_latitude"}
    )

    lat_lon_lookup = build_candidate_lookup(
        lat_lon_candidate,
        key_col="__row_id__",
        latitude_col="point_latitude",
        longitude_col="point_longitude",
        nuts_gdf=nuts_gdf,
    ).set_index("__row_id__").reindex(result.index)
    lon_lat_lookup = build_candidate_lookup(
        lon_lat_candidate,
        key_col="__row_id__",
        latitude_col="point_latitude",
        longitude_col="point_longitude",
        nuts_gdf=nuts_gdf,
    ).set_index("__row_id__").reindex(result.index)

    lat_lon_score = lat_lon_lookup["nuts_match_level_count"].fillna(0).astype(int)
    lon_lat_score = lon_lat_lookup["nuts_match_level_count"].fillna(0).astype(int)

    country_bounds = tuple(float(value) for value in nuts_gdf.total_bounds)
    lat_lon_bbox = is_within_country_bounds(result["__coord_1__"], result["__coord_2__"], country_bounds)
    lon_lat_bbox = is_within_country_bounds(result["__coord_2__"], result["__coord_1__"], country_bounds)

    id_adr_text = result[id_adr_col].astype("string").str.strip()
    missing_id_adr = result[id_adr_col].isna() | id_adr_text.isna() | id_adr_text.eq("")
    parse_failed = ~missing_id_adr & (result["__coord_1__"].isna() | result["__coord_2__"].isna())

    use_lat_lon_from_nuts = lat_lon_score.gt(lon_lat_score)
    use_lon_lat_from_nuts = lon_lat_score.gt(lat_lon_score)
    tied_scores = lat_lon_score.eq(lon_lat_score)
    use_lat_lon_from_bbox = tied_scores & lat_lon_bbox & ~lon_lat_bbox
    use_lon_lat_from_bbox = tied_scores & lon_lat_bbox & ~lat_lon_bbox
    unresolved = ~(use_lat_lon_from_nuts | use_lon_lat_from_nuts | use_lat_lon_from_bbox | use_lon_lat_from_bbox | missing_id_adr | parse_failed)

    choose_lat_lon = use_lat_lon_from_nuts | use_lat_lon_from_bbox
    choose_lon_lat = use_lon_lat_from_nuts | use_lon_lat_from_bbox
    if default_coordinate_order == "lat_lon":
        choose_lat_lon = choose_lat_lon | unresolved
    else:
        choose_lon_lat = choose_lon_lat | unresolved

    result["point_latitude"] = np.nan
    result["point_longitude"] = np.nan
    result.loc[choose_lat_lon, "point_latitude"] = result.loc[choose_lat_lon, "__coord_1__"]
    result.loc[choose_lat_lon, "point_longitude"] = result.loc[choose_lat_lon, "__coord_2__"]
    result.loc[choose_lon_lat, "point_latitude"] = result.loc[choose_lon_lat, "__coord_2__"]
    result.loc[choose_lon_lat, "point_longitude"] = result.loc[choose_lon_lat, "__coord_1__"]

    result["id_adr_coordinate_order"] = pd.NA
    result.loc[choose_lat_lon, "id_adr_coordinate_order"] = "lat_lon"
    result.loc[choose_lon_lat, "id_adr_coordinate_order"] = "lon_lat"
    result.loc[missing_id_adr, "id_adr_coordinate_order"] = "missing"
    result.loc[parse_failed, "id_adr_coordinate_order"] = "unparsed"

    result["id_adr_order_resolution"] = pd.NA
    result.loc[use_lat_lon_from_nuts | use_lon_lat_from_nuts, "id_adr_order_resolution"] = "nuts_match"
    result.loc[use_lat_lon_from_bbox | use_lon_lat_from_bbox, "id_adr_order_resolution"] = "country_bbox"
    result.loc[unresolved & ~missing_id_adr & ~parse_failed, "id_adr_order_resolution"] = "default_order"
    result.loc[missing_id_adr, "id_adr_order_resolution"] = "missing_id_adr"
    result.loc[parse_failed, "id_adr_order_resolution"] = "parse_failed"

    for column in ["nuts1_code", "nuts1_name", "nuts2_code", "nuts2_name", "nuts3_code", "nuts3_name"]:
        result[column] = pd.NA
        result.loc[choose_lat_lon, column] = lat_lon_lookup.loc[choose_lat_lon, column]
        result.loc[choose_lon_lat, column] = lon_lat_lookup.loc[choose_lon_lat, column]

    result = result.drop(columns=["__row_id__", "__coord_1__", "__coord_2__"], errors="ignore")
    return reorder_enrichment_columns(result, id_adr_col=id_adr_col)


def write_updated_excel_sheet(
    workbook_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} was not found in {workbook_path.name}.")

    sheet_index = workbook.sheetnames.index(sheet_name)
    del workbook[sheet_name]
    worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([normalize_excel_cell(value) for value in row])
    style_worksheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_updated_output(
    export_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    df: pd.DataFrame,
    verbose: bool,
) -> None:
    suffix = export_path.suffix.lower()
    if suffix == ".csv":
        separator = detect_csv_separator(export_path)
        log_progress(f"Writing updated csv to {output_path}...", enabled=verbose)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output_path, index=False, encoding="utf-8-sig", sep=separator)
        return
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        log_progress(f"Writing updated workbook to {output_path}...", enabled=verbose)
        write_updated_excel_sheet(export_path, output_path, sheet_name=sheet_name, df=df)
        return
    raise ValueError(f"Unsupported FLOOD_LGD file type: {export_path.suffix}")


def run(args: argparse.Namespace) -> None:
    verbose = not args.quiet

    export_path = Path(args.flood_lgd_file)
    nuts_path = Path(args.nuts_file)

    ensure_required_file(export_path, "FLOOD_LGD file")
    ensure_required_file(nuts_path, "NUTS file")

    if args.in_place and args.output_file:
        raise ValueError("Use either --in-place or --output-file, not both.")
    output_path = export_path if args.in_place else Path(args.output_file) if args.output_file else derive_output_path(export_path)

    export_df = read_flood_lgd_frame(export_path, sheet_name=args.sheet_name, verbose=verbose)
    log_progress(f"Loading NUTS polygons from {nuts_path} for country {args.country_code}...", enabled=verbose)
    nuts_gdf = load_nuts(str(nuts_path), target_countries={str(args.country_code).strip().upper()})

    updated_df = attach_nuts_metadata(
        export_df,
        nuts_gdf,
        id_adr_col=args.id_adr_col,
        default_coordinate_order=args.default_coordinate_order,
    )

    matched_nuts3 = int(updated_df["nuts3_code"].notna().sum()) if "nuts3_code" in updated_df.columns else 0
    log_progress(
        f"Prepared updated FLOOD_LGD table with {matched_nuts3:,} rows carrying a NUTS 3 assignment.",
        enabled=verbose,
    )
    write_updated_output(
        export_path,
        output_path,
        sheet_name=args.sheet_name,
        df=updated_df,
        verbose=verbose,
    )
    print(f"Done. Output written to: {output_path.resolve()}")


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
