from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from build_flood_lgd_exports import (
    DEFAULT_CSV_SEPARATOR,
    DEFAULT_SHEET_NAME,
    ensure_required_file,
    load_source_frame,
    log_progress,
    normalize_excel_cell,
    normalize_point_id_series,
    style_worksheet,
)


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Add the Default_Date metadata column from the original T20 workbook "
            "to an already-built FLOOD_LGD csv/xlsx without recomputing flood merging."
        )
    )
    parser.add_argument("--source-workbook", required=True, help="Original T20 workbook used to recover Default_Date by point_id.")
    parser.add_argument("--flood-lgd-file", required=True, help="Existing FLOOD_LGD csv or xlsx file to enrich.")
    parser.add_argument("--output-file", default=None, help="Optional output path. Default writes a sibling file with a _with_default_date suffix.")
    parser.add_argument("--in-place", action="store_true", help="Overwrite --flood-lgd-file instead of writing a sibling file.")
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="Sheet name to update when --flood-lgd-file is an Excel workbook.")
    parser.add_argument("--source-sheet-name", default=None, help="Optional source workbook sheet name. Default uses the first sheet.")
    parser.add_argument("--source-point-id-col", default=None, help="Optional source point identifier column override. Leave blank to auto-detect.")
    parser.add_argument("--source-default-date-col", default=None, help="Optional source Default_Date column override. Leave blank to auto-detect.")
    parser.add_argument("--quiet", action="store_true", help="Disable progress logging and only print the final completion message.")
    return parser


def derive_output_path(export_path: Path) -> Path:
    return export_path.with_name(f"{export_path.stem}_with_default_date{export_path.suffix}")


def read_flood_lgd_frame(export_path: Path, *, sheet_name: str, verbose: bool) -> pd.DataFrame:
    suffix = export_path.suffix.lower()
    log_progress(f"Loading existing FLOOD_LGD output from {export_path}...", enabled=verbose)
    if suffix == ".csv":
        return pd.read_csv(export_path, sep=DEFAULT_CSV_SEPARATOR)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return pd.read_excel(export_path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported FLOOD_LGD file type: {export_path.suffix}")


def reorder_default_date_column(df: pd.DataFrame) -> pd.DataFrame:
    if "Default_Date" not in df.columns:
        return df

    ordered_columns = [column for column in df.columns if column != "Default_Date"]
    if "CLOSED_DEFAULT_DATE" in ordered_columns:
        insert_at = ordered_columns.index("CLOSED_DEFAULT_DATE") + 1
    elif "Facility_ID" in ordered_columns:
        insert_at = ordered_columns.index("Facility_ID") + 1
    elif "point_id" in ordered_columns:
        insert_at = ordered_columns.index("point_id") + 1
    else:
        insert_at = len(ordered_columns)
    ordered_columns.insert(insert_at, "Default_Date")
    return df[ordered_columns].copy()


def attach_default_date(export_df: pd.DataFrame, source_df: pd.DataFrame) -> pd.DataFrame:
    if "point_id" not in export_df.columns:
        raise KeyError("The FLOOD_LGD file must contain a point_id column.")

    result = export_df.copy()
    result["point_id"] = normalize_point_id_series(result["point_id"])

    source_lookup = source_df.copy()
    source_lookup["point_id"] = normalize_point_id_series(source_lookup["point_id"])
    source_lookup = source_lookup[["point_id", "Default_Date"]].drop_duplicates(subset=["point_id"], keep="first")

    result = result.drop(columns=["Default_Date"], errors="ignore")
    result = result.merge(source_lookup, on="point_id", how="left")
    return reorder_default_date_column(result)


def write_updated_excel_sheet(
    workbook_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    df: pd.DataFrame,
) -> None:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} was not found in {workbook_path.name}.")

    sheet_index = workbook.sheetnames.index(sheet_name)
    del workbook[sheet_name]
    worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    worksheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        worksheet.append([normalize_excel_cell(value) for value in row])
    style_worksheet(worksheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_updated_output(
    export_path: Path,
    output_path: Path,
    *,
    sheet_name: str,
    df: pd.DataFrame,
    verbose: bool,
) -> None:
    suffix = export_path.suffix.lower()
    if suffix == ".csv":
        log_progress(f"Writing updated csv to {output_path}...", enabled=verbose)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output_path, index=False, encoding="utf-8", sep=DEFAULT_CSV_SEPARATOR)
        return
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        log_progress(f"Writing updated workbook to {output_path}...", enabled=verbose)
        write_updated_excel_sheet(export_path, output_path, sheet_name=sheet_name, df=df)
        return
    raise ValueError(f"Unsupported FLOOD_LGD file type: {export_path.suffix}")


def run(args: argparse.Namespace) -> None:
    verbose = not args.quiet

    source_workbook = Path(args.source_workbook)
    export_path = Path(args.flood_lgd_file)

    ensure_required_file(source_workbook, "source workbook")
    ensure_required_file(export_path, "FLOOD_LGD file")

    if args.in_place and args.output_file:
        raise ValueError("Use either --in-place or --output-file, not both.")
    output_path = export_path if args.in_place else Path(args.output_file) if args.output_file else derive_output_path(export_path)

    source_df = load_source_frame(
        source_workbook,
        sheet_name=args.source_sheet_name,
        point_id_col=args.source_point_id_col,
        default_date_col=args.source_default_date_col,
    )
    export_df = read_flood_lgd_frame(export_path, sheet_name=args.sheet_name, verbose=verbose)
    updated_df = attach_default_date(export_df, source_df)

    populated_count = int(updated_df["Default_Date"].notna().sum()) if "Default_Date" in updated_df.columns else 0
    log_progress(
        f"Prepared updated FLOOD_LGD table with {populated_count:,} rows carrying Default_Date values.",
        enabled=verbose,
    )
    write_updated_output(
        export_path,
        output_path,
        sheet_name=args.sheet_name,
        df=updated_df,
        verbose=verbose,
    )
    print(f"Done. Output written to: {output_path.resolve()}")


def main() -> None:
    parser = build_argument_parser()
    args = parser.parse_args()
    run(args)


if __name__ == "__main__":
    main()
