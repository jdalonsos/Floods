from __future__ import annotations

from pathlib import Path
import sys
import tempfile
import unittest

import numpy as np
from openpyxl import load_workbook
import pandas as pd
import rasterio
from rasterio.transform import from_origin

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
SITE_PACKAGES = PROJECT_ROOT / ".venv" / "Lib" / "site-packages"
if str(SITE_PACKAGES) not in sys.path:
    sys.path.insert(0, str(SITE_PACKAGES))
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from check_points_against_jrc_floods import (  # noqa: E402
    GASPAR_EVENT_HITS_COLUMNS,
    JRC_EVENT_HITS_COLUMNS,
    LONGITUDE_ALIASES,
    PointColumns,
    build_candidate_sheet,
    build_detailed_sheet,
    build_gaspar_candidate_sheet,
    build_gaspar_hits_sheet,
    build_hanze_candidate_sheet,
    build_hanze_hits_sheet,
    build_point_flag_sheet,
    build_hits_sheet,
    build_summary_table,
    build_row_level_study_periods,
    compute_buffer_stats,
    filter_candidate_events_by_row_study_period,
    list_tri_inondable_members,
    parse_coordinate_series,
    resolve_named_column,
    split_dataframe_for_excel,
    write_gaspar_output_workbook,
)
from check_points_against_jrc_floods_collaterals import (  # noqa: E402
    build_collaterals_argument_parser,
)


class CheckPointsAgainstJrcFloodsTests(unittest.TestCase):
    def test_collaterals_variant_parser_uses_expected_defaults(self) -> None:
        parser = build_collaterals_argument_parser()

        args = parser.parse_args([])

        self.assertEqual(args.latitude_col, "lat")
        self.assertEqual(args.longitude_col, "lon")
        self.assertEqual(args.point_id_col, "ID_geoloc")
        self.assertEqual(args.row_study_end_col, "last_date")
        self.assertEqual(args.study_start, "2000-01-01")
        self.assertEqual(
            Path(args.out_file),
            Path("data/processed/france_points_jrc_flood_check_collaterals.xlsx"),
        )

    def test_longitude_aliases_accept_long(self) -> None:
        df = pd.DataFrame(columns=["LAT", "LONG"])
        self.assertEqual(resolve_named_column(df, None, LONGITUDE_ALIASES), "LONG")

    def test_parse_coordinate_series_accepts_decimal_commas(self) -> None:
        series = pd.Series(["47,87431063", "-2,13106221", "43,70053124", pd.NA])
        parsed = parse_coordinate_series(series)

        self.assertAlmostEqual(float(parsed.iloc[0]), 47.87431063)
        self.assertAlmostEqual(float(parsed.iloc[1]), -2.13106221)
        self.assertAlmostEqual(float(parsed.iloc[2]), 43.70053124)
        self.assertTrue(pd.isna(parsed.iloc[3]))

    def test_parse_coordinate_series_handles_mixed_decimal_formats(self) -> None:
        series = pd.Series(["48.87292437", "2,39401056", "1,234.56", "1.234,56"])
        parsed = parse_coordinate_series(series)

        self.assertAlmostEqual(float(parsed.iloc[0]), 48.87292437)
        self.assertAlmostEqual(float(parsed.iloc[1]), 2.39401056)
        self.assertAlmostEqual(float(parsed.iloc[2]), 1234.56)
        self.assertAlmostEqual(float(parsed.iloc[3]), 1234.56)

    def test_list_tri_inondable_members_accepts_directory_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tri_dir = Path(tmp_dir)
            (tri_dir / "n_inondable_01_01for_s.shp").touch()
            (tri_dir / "n_inondable_03_03mcc_s.shp").touch()
            (tri_dir / "n_commune_s.shp").touch()

            members = list_tri_inondable_members(tri_dir)

        self.assertEqual(
            members,
            [
                ("n_inondable_01_01for_s.shp", "01for"),
                ("n_inondable_03_03mcc_s.shp", "03mcc"),
            ],
        )

    def test_compute_buffer_stats_uses_full_buffer_pixels_for_percentage(self) -> None:
        data = np.zeros((5, 5), dtype=np.float32)
        data[2, 2] = 10.0
        data[1, 2] = 20.0

        with tempfile.TemporaryDirectory() as tmpdir:
            raster_path = Path(tmpdir) / "buffer_test.tif"
            with rasterio.open(
                raster_path,
                "w",
                driver="GTiff",
                height=5,
                width=5,
                count=1,
                dtype="float32",
                crs="EPSG:3857",
                transform=from_origin(0, 100, 20, 20),
                nodata=0.0,
            ) as dst:
                dst.write(data, 1)

            with rasterio.open(raster_path) as src:
                stats = compute_buffer_stats(
                    src,
                    x=50.0,
                    y=50.0,
                    radius_m=100.0,
                    threshold_cm=0.0,
                    prefix="buffer",
                )

        self.assertEqual(int(stats["buffer_total_pixels"]), 25)
        self.assertEqual(int(stats["buffer_flooded_pixels"]), 2)
        self.assertAlmostEqual(float(stats["buffer_flooded_pixel_pct"]), 8.0)
        self.assertEqual(float(stats["buffer_min_depth_cm"]), 10.0)
        self.assertEqual(float(stats["buffer_max_depth_cm"]), 20.0)

    def test_build_row_level_study_periods_uses_full_history_and_fallback_end_date(self) -> None:
        points_df = pd.DataFrame(
            {
                "Reference_Date": ["31/12/2008", "31/12/2011"],
                "Closed_Default_Date": ["08/10/2013", pd.NA],
                "Cut_off_Date": ["31/12/2024", "31/12/2024"],
            }
        )

        result, resolved = build_row_level_study_periods(
            points_df,
            anchor_col="Reference_Date",
            end_col="Closed_Default_Date",
            fallback_end_col="Cut_off_Date",
            lookback_years=None,
        )

        self.assertIsNotNone(resolved)
        assert resolved is not None
        self.assertEqual(resolved.anchor, "Reference_Date")
        self.assertEqual(resolved.primary_end, "Closed_Default_Date")
        self.assertEqual(resolved.fallback_end, "Cut_off_Date")
        self.assertTrue(pd.isna(result.loc[0, "study_period_start"]))
        self.assertEqual(result.loc[0, "study_period_end"], pd.Timestamp("2013-10-08"))
        self.assertEqual(result.loc[0, "study_period_end_source"], "Closed_Default_Date")
        self.assertTrue(pd.isna(result.loc[1, "study_period_start"]))
        self.assertEqual(result.loc[1, "study_period_end"], pd.Timestamp("2024-12-31"))
        self.assertEqual(result.loc[1, "study_period_end_source"], "Cut_off_Date")

    def test_build_row_level_study_periods_supports_optional_lookback_years(self) -> None:
        points_df = pd.DataFrame(
            {
                "Reference_Date": ["31/12/2008"],
                "Closed_Default_Date": ["08/10/2013"],
                "Cut_off_Date": ["31/12/2024"],
            }
        )

        result, _ = build_row_level_study_periods(
            points_df,
            anchor_col="Reference_Date",
            end_col="Closed_Default_Date",
            fallback_end_col="Cut_off_Date",
            lookback_years=5,
        )

        self.assertEqual(result.loc[0, "study_period_start"], pd.Timestamp("2003-12-31"))

    def test_filter_candidate_events_by_row_study_period_is_point_specific(self) -> None:
        candidate_df = pd.DataFrame(
            {
                "point_id": [1, 1, 1, 2, 3],
                "event_id": ["too_early", "overlap", "too_late", "fallback_match", pd.NA],
                "start_date": pd.to_datetime(
                    ["2001-01-01", "2008-02-01", "2014-01-01", "2020-05-01", pd.NaT]
                ),
                "end_date": pd.to_datetime(
                    ["2002-01-01", "2008-03-01", "2014-02-01", "2020-05-20", pd.NaT]
                ),
                "study_period_start": pd.to_datetime(
                    [pd.NaT, pd.NaT, pd.NaT, pd.NaT, pd.NaT]
                ),
                "study_period_end": pd.to_datetime(
                    ["2013-10-08", "2013-10-08", "2013-10-08", "2024-12-31", pd.NaT]
                ),
            }
        )

        filtered = filter_candidate_events_by_row_study_period(candidate_df)

        self.assertEqual(
            filtered["event_id"].fillna("missing").tolist(),
            ["too_early", "overlap", "fallback_match", "missing"],
        )

    def test_build_summary_table_handles_no_positive_hits(self) -> None:
        original_points = pd.DataFrame(
            {
                "point_id": [1],
                "LAT": [48.8566],
                "LONG": [2.3522],
            }
        )
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": ["event-1"],
                "start_date": pd.to_datetime(["2020-01-01"]),
                "end_date": pd.to_datetime(["2020-01-03"]),
                "raster_path_found": [True],
            }
        )
        inspected_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": ["event-1"],
                "hit_at_point": [False],
                "exact_point_depth_cm": [pd.NA],
                "point_buffer_flood_hit": [False],
                "point_buffer_flooded_pixels": [0],
                "point_buffer_flooded_area_m2": [0.0],
                "point_buffer_max_depth_cm": [pd.NA],
                "point_buffer_median_depth_cm": [pd.NA],
                "point_buffer_mean_depth_cm": [pd.NA],
                "buffer_flood_hit": [False],
                "buffer_flooded_pixels": [0],
                "buffer_flooded_area_m2": [0.0],
                "buffer_max_depth_cm": [pd.NA],
                "buffer_median_depth_cm": [pd.NA],
                "buffer_mean_depth_cm": [pd.NA],
                "surrounding_buffer_flood_hit": [False],
                "surrounding_buffer_flooded_pixels": [0],
                "surrounding_buffer_flooded_area_m2": [0.0],
                "surrounding_buffer_max_depth_cm": [pd.NA],
                "surrounding_buffer_median_depth_cm": [pd.NA],
                "surrounding_buffer_mean_depth_cm": [pd.NA],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=inspected_df,
            default_date_inspected_df=None,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
        )

        self.assertFalse(bool(summary.loc[0, "jrc_flood_hit"]))
        self.assertEqual(summary.loc[0, "jrc_flood_flag"], "no")
        self.assertEqual(summary.loc[0, "point_buffer_radius_m"], 40.0)
        self.assertEqual(summary.loc[0, "buffer_radius_km"], 1.0)

    def test_build_summary_table_separates_point_and_surrounding_buffer_metrics(self) -> None:
        original_points = pd.DataFrame(
            {
                "point_id": [1],
                "LAT": [48.8566],
                "LONG": [2.3522],
                "study_period_anchor_date": pd.to_datetime(["2020-01-15"]),
                "study_period_end": pd.to_datetime(["2020-01-31"]),
                "study_period_fallback_end_date": pd.to_datetime(["2020-03-31"]),
            }
        )
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1, 1, 1],
                "event_id": ["event-1", "event-2", "event-3"],
                "start_date": pd.to_datetime(["2020-01-01", "2020-01-14", "2020-01-20"]),
                "end_date": pd.to_datetime(["2020-01-03", "2020-01-16", "2020-01-22"]),
                "raster_path_found": [True, True, True],
            }
        )
        inspected_df = pd.DataFrame(
            {
                "point_id": [1, 1, 1],
                "event_id": ["event-1", "event-2", "event-3"],
                "hit_at_point": [False, True, True],
                "exact_point_depth_cm": [30.0, 45.0, 50.0],
                "point_buffer_flood_hit": [False, True, True],
                "point_buffer_flooded_pixels": [0, 2, 3],
                "point_buffer_flooded_area_m2": [0.0, 800.0, 1200.0],
                "point_buffer_max_depth_cm": [pd.NA, 45.0, 50.0],
                "point_buffer_median_depth_cm": [pd.NA, 25.0, 30.0],
                "point_buffer_mean_depth_cm": [pd.NA, 27.0, 35.0],
                "buffer_flood_hit": [True, True, True],
                "buffer_flooded_pixels": [10, 12, 15],
                "buffer_flooded_area_m2": [4000.0, 4800.0, 6000.0],
                "buffer_max_depth_cm": [30.0, 55.0, 60.0],
                "buffer_median_depth_cm": [20.0, 22.0, 25.0],
                "buffer_mean_depth_cm": [18.0, 24.0, 28.0],
                "surrounding_buffer_flood_hit": [True, True, True],
                "surrounding_buffer_flooded_pixels": [10, 12, 15],
                "surrounding_buffer_flooded_area_m2": [4000.0, 4800.0, 6000.0],
                "surrounding_buffer_max_depth_cm": [30.0, 55.0, 60.0],
                "surrounding_buffer_median_depth_cm": [20.0, 22.0, 25.0],
                "surrounding_buffer_mean_depth_cm": [18.0, 24.0, 28.0],
            }
        )
        default_date_inspected_df = pd.DataFrame(
            {
                "point_id": [1, 1],
                "event_id": ["event-1", "event-2"],
                "hit_at_point": [False, True],
                "exact_point_depth_cm": [30.0, 50.0],
                "point_buffer_flood_hit": [False, True],
                "point_buffer_flooded_pixels": [0, 3],
                "point_buffer_flooded_area_m2": [0.0, 1200.0],
                "point_buffer_max_depth_cm": [pd.NA, 50.0],
                "point_buffer_median_depth_cm": [pd.NA, 30.0],
                "point_buffer_mean_depth_cm": [pd.NA, 35.0],
                "buffer_flood_hit": [True, True],
                "buffer_flooded_pixels": [10, 15],
                "buffer_flooded_area_m2": [4000.0, 6000.0],
                "buffer_max_depth_cm": [30.0, 60.0],
                "buffer_median_depth_cm": [20.0, 25.0],
                "buffer_mean_depth_cm": [18.0, 28.0],
                "surrounding_buffer_flood_hit": [True, True],
                "surrounding_buffer_flooded_pixels": [10, 15],
                "surrounding_buffer_flooded_area_m2": [4000.0, 6000.0],
                "surrounding_buffer_max_depth_cm": [30.0, 60.0],
                "surrounding_buffer_median_depth_cm": [20.0, 25.0],
                "surrounding_buffer_mean_depth_cm": [18.0, 28.0],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=inspected_df,
            default_date_inspected_df=default_date_inspected_df,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
        )

        self.assertEqual(int(summary.loc[0, "hit_at_point_event_count"]), 2)
        self.assertEqual(int(summary.loc[0, "hit_within_buffer_event_count"]), 3)
        self.assertEqual(int(summary.loc[0, "hit_event_count"]), 3)
        self.assertEqual(int(summary.loc[0, "hit_event_count_until_default_date"]), 2)
        self.assertEqual(float(summary.loc[0, "max_point_buffer_depth_cm"]), 50.0)
        self.assertEqual(float(summary.loc[0, "max_buffer_depth_cm"]), 60.0)
        self.assertEqual(int(summary.loc[0, "max_buffer_flooded_pixels"]), 15)
        self.assertTrue(bool(summary.loc[0, "jrc_flood_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_jrc"]), 1)
        self.assertEqual(int(summary.loc[0, "flag_gaspar"]), 0)
        self.assertEqual(int(summary.loc[0, "flag_flood"]), 1)
        self.assertEqual(summary.loc[0, "flag_flood_source"], "jrc")
        self.assertEqual(summary.loc[0, "flag_flood_case"], "case_a_jrc")

    def test_candidate_and_hits_sheets_include_min_depth_and_flooded_pixel_percentage(self) -> None:
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "LAT": [48.8566],
                "LONG": [2.3522],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_name": ["Paris"],
                "event_id": ["event-1"],
                "raster_file": ["flood.tif"],
                "resolved_raster_path": ["D:/fake/flood.tif"],
                "start_date": pd.to_datetime(["2020-01-01"]),
                "end_date": pd.to_datetime(["2020-01-03"]),
                "duration_days": [3],
                "max_depth_cm": [90.0],
                "flooded_pixels": [100],
                "flooded_area_m2": [40000.0],
                "raster_path_found": [True],
            }
        )
        inspected_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": ["event-1"],
                "hit_at_point": [True],
                "exact_point_depth_cm": [50.0],
                "point_buffer_total_pixels": [20],
                "point_buffer_flood_hit": [True],
                "point_buffer_flooded_pixels": [5],
                "point_buffer_flooded_pixel_pct": [25.0],
                "point_buffer_flooded_area_m2": [2000.0],
                "point_buffer_min_depth_cm": [10.0],
                "point_buffer_max_depth_cm": [50.0],
                "point_buffer_median_depth_cm": [20.0],
                "point_buffer_mean_depth_cm": [25.0],
                "point_buffer_radius_m": [40.0],
                "buffer_total_pixels": [100],
                "buffer_flood_hit": [True],
                "buffer_flooded_pixels": [12],
                "buffer_flooded_pixel_pct": [12.0],
                "buffer_flooded_area_m2": [4800.0],
                "buffer_min_depth_cm": [5.0],
                "buffer_max_depth_cm": [60.0],
                "buffer_median_depth_cm": [25.0],
                "buffer_mean_depth_cm": [28.0],
                "buffer_radius_km": [1.0],
                "surrounding_buffer_total_pixels": [100],
                "surrounding_buffer_flood_hit": [True],
                "surrounding_buffer_flooded_pixels": [12],
                "surrounding_buffer_flooded_pixel_pct": [12.0],
                "surrounding_buffer_flooded_area_m2": [4800.0],
                "surrounding_buffer_min_depth_cm": [5.0],
                "surrounding_buffer_max_depth_cm": [60.0],
                "surrounding_buffer_median_depth_cm": [25.0],
                "surrounding_buffer_mean_depth_cm": [28.0],
                "surrounding_buffer_radius_km": [1.0],
            }
        )

        candidate_sheet = build_candidate_sheet(
            candidate_df=candidate_df,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            inspected_df=inspected_df,
            row_study_period_columns=None,
        )
        hits_sheet = build_hits_sheet(candidate_sheet)

        self.assertEqual(float(candidate_sheet.loc[0, "point_buffer_min_depth_cm"]), 10.0)
        self.assertEqual(float(candidate_sheet.loc[0, "buffer_min_depth_cm"]), 5.0)
        self.assertEqual(float(candidate_sheet.loc[0, "surrounding_buffer_min_depth_cm"]), 5.0)
        self.assertEqual(float(candidate_sheet.loc[0, "buffer_flooded_pixel_pct"]), 12.0)
        self.assertEqual(float(candidate_sheet.loc[0, "surrounding_buffer_flooded_pixel_pct"]), 12.0)
        self.assertEqual(int(candidate_sheet.loc[0, "buffer_total_pixels"]), 100)
        self.assertEqual(len(hits_sheet), 1)
        expected_hits_columns = [column for column in JRC_EVENT_HITS_COLUMNS if column in candidate_sheet.columns]
        self.assertEqual(hits_sheet.columns.tolist(), expected_hits_columns)
        self.assertIn("buffer_min_depth_cm", hits_sheet.columns)
        self.assertIn("buffer_flooded_pixel_pct", hits_sheet.columns)
        self.assertNotIn("resolved_raster_path", hits_sheet.columns)
        self.assertNotIn("surrounding_buffer_total_pixels", hits_sheet.columns)

    def test_build_point_flag_sheet_marks_only_hit_points(self) -> None:
        points_df = pd.DataFrame({"point_id": [1, 2, 3, 4]})

        result = build_point_flag_sheet(points_df, "point_id", {2, 4})

        self.assertEqual(result["point_id"].tolist(), [1, 2, 3, 4])
        self.assertEqual(result["flag_flood"].tolist(), [0, 1, 0, 1])

    def test_build_detailed_sheet_keeps_original_rows_and_prepends_touch_flag(self) -> None:
        points_df = pd.DataFrame(
            {
                "point_id": [11, 12],
                "LAT": [47.5, 48.5],
                "LONG": [6.8, 2.3],
                "Reference_Date": pd.to_datetime(["2020-01-01", "2021-01-01"]),
            }
        )

        result = build_detailed_sheet(points_df, "point_id", {12})

        self.assertEqual(
            result.columns.tolist(),
            ["point_id", "touched", "LAT", "LONG", "Reference_Date"],
        )
        self.assertEqual(result["point_id"].tolist(), [11, 12])
        self.assertEqual(result["touched"].tolist(), [0, 1])
        self.assertEqual(result["LAT"].tolist(), [47.5, 48.5])

    def test_split_dataframe_for_excel_splits_large_sheets_into_numbered_tabs(self) -> None:
        df = pd.DataFrame({"value": [1, 2, 3, 4, 5]})

        parts = split_dataframe_for_excel(df, "candidate_events", max_rows=2)

        self.assertEqual([sheet_name for sheet_name, _ in parts], ["candidate_events", "candidate_events_2", "candidate_events_3"])
        self.assertEqual([len(chunk) for _, chunk in parts], [2, 2, 1])
        self.assertEqual(parts[1][1]["value"].tolist(), [3, 4])

    def test_write_gaspar_output_workbook_creates_expected_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "gaspar_output.xlsx"

            write_gaspar_output_workbook(
                output_path=output_path,
                point_flag_sheet=pd.DataFrame({"point_id": [1], "flag_flood": [1]}),
                detailed_sheet=pd.DataFrame({"point_id": [1], "touched": [1], "LAT": [45.0]}),
                candidate_sheet=pd.DataFrame({"point_id": [1], "gaspar_event_uid": ["g1"]}),
                hits_sheet=pd.DataFrame({"point_id": [1], "gaspar_event_uid": ["g1"]}),
            )

            workbook = load_workbook(output_path, read_only=True)

            self.assertEqual(
                workbook.sheetnames,
                ["point_flags", "Detailed", "candidate_events", "event_hits"],
            )
            point_flags_rows = list(workbook["point_flags"].iter_rows(values_only=True))
            self.assertEqual(point_flags_rows[0], ("point_id", "flag_flood"))
            self.assertEqual(point_flags_rows[1], (1, 1))
            workbook.close()

    def test_gaspar_candidate_and_hits_sheets_keep_only_tri_for_or_riparian(self) -> None:
        gaspar_candidate_df = pd.DataFrame(
            {
                "point_id": [1, 2, 3],
                "LAT": [45.0, 46.0, 47.0],
                "LONG": [3.0, 4.0, 5.0],
                "excel_row_number": [2, 3, 4],
                "lau_code": ["FR_A", "FR_B", "FR_C"],
                "lau_name": ["A", "B", "C"],
                "insee_com": ["00001", "00002", "00003"],
                "study_period_end": pd.to_datetime(["2024-12-31", "2024-12-31", "2024-12-31"]),
                "gaspar_event_uid": ["g1", "g2", "g3"],
                "cod_nat_catnat": ["c1", "c2", "c3"],
                "gaspar_start_date": pd.to_datetime(["2020-01-01", "2020-02-01", "2020-03-01"]),
                "gaspar_end_date": pd.to_datetime(["2020-01-03", "2020-02-03", "2020-03-03"]),
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1, 2, 3],
                "tri_for_hit": [True, False, False],
                "tri_boundary_hit": [True, True, False],
                "tri_zone_status": ["for", "inside_n_tri_not_for", "outside_n_tri"],
                "riparian_hit": [False, False, True],
            }
        )

        candidate_sheet = build_gaspar_candidate_sheet(
            gaspar_candidate_df=gaspar_candidate_df,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            tri_classification_df=tri_classification_df,
            row_study_period_columns=None,
        )
        hits_sheet = build_gaspar_hits_sheet(candidate_sheet)

        self.assertEqual(candidate_sheet["gaspar_spatial_hit"].tolist(), [True, False, True])
        self.assertEqual(candidate_sheet["gaspar_hit_reason"].tolist(), ["tri_for", "not_selected", "riparian_outside_n_tri"])
        self.assertEqual(hits_sheet["point_id"].tolist(), [1, 3])
        self.assertEqual(hits_sheet["gaspar_event_uid"].tolist(), ["g1", "g3"])
        expected_hits_columns = [column for column in GASPAR_EVENT_HITS_COLUMNS if column in candidate_sheet.columns]
        self.assertEqual(hits_sheet.columns.tolist(), expected_hits_columns)
        self.assertNotIn("gaspar_spatial_hit", hits_sheet.columns)

    def test_build_summary_table_uses_gaspar_tri_for_when_jrc_is_negative(self) -> None:
        original_points = pd.DataFrame({"point_id": [1], "LAT": [48.8566], "LONG": [2.3522]})
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
                "insee_com": ["75056"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": [pd.NA],
                "start_date": pd.to_datetime([pd.NaT]),
                "end_date": pd.to_datetime([pd.NaT]),
                "raster_path_found": [False],
            }
        )
        inspected_df = pd.DataFrame()
        gaspar_candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "gaspar_event_uid": ["gaspar-1"],
                "cod_nat_catnat": ["CATNAT-1"],
                "gaspar_start_date": pd.to_datetime(["2021-05-10"]),
                "gaspar_end_date": pd.to_datetime(["2021-05-12"]),
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1],
                "tri_for_hit": [True],
                "tri_boundary_hit": [True],
                "tri_zone_status": ["for"],
                "riparian_hit": [False],
                "tri_scenario_labels": ["Aléa de forte probabilité"],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=inspected_df,
            default_date_inspected_df=None,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
            gaspar_candidate_df=gaspar_candidate_df,
            tri_classification_df=tri_classification_df,
        )

        self.assertFalse(bool(summary.loc[0, "jrc_flood_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_jrc"]), 0)
        self.assertTrue(bool(summary.loc[0, "gaspar_commune_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_gaspar"]), 1)
        self.assertTrue(bool(summary.loc[0, "tri_for_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_flood"]), 1)
        self.assertEqual(summary.loc[0, "flag_flood_source"], "gaspar")
        self.assertEqual(summary.loc[0, "flag_flood_case"], "case_b_gaspar_tri_for")
        self.assertEqual(summary.loc[0, "flag_flood_start_date"], pd.Timestamp("2021-05-10"))
        self.assertEqual(summary.loc[0, "flag_flood_end_date"], pd.Timestamp("2021-05-12"))

    def test_build_summary_table_keeps_gaspar_inside_n_tri_not_for_negative(self) -> None:
        original_points = pd.DataFrame({"point_id": [1], "LAT": [48.8566], "LONG": [2.3522]})
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
                "insee_com": ["75056"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": [pd.NA],
                "start_date": pd.to_datetime([pd.NaT]),
                "end_date": pd.to_datetime([pd.NaT]),
                "raster_path_found": [False],
            }
        )
        gaspar_candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "gaspar_event_uid": ["gaspar-2"],
                "cod_nat_catnat": ["CATNAT-2"],
                "gaspar_start_date": pd.to_datetime(["2019-11-01"]),
                "gaspar_end_date": pd.to_datetime(["2019-11-04"]),
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1],
                "tri_for_hit": [False],
                "tri_boundary_hit": [True],
                "tri_zone_status": ["inside_n_tri_not_for"],
                "riparian_hit": [False],
                "flood_risk_area_value": ["out"],
                "TRI": ["out"],
                "tri_scenario_codes": [pd.NA],
                "tri_scenario_labels": [pd.NA],
                "riparian_zone_hit": [False],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=pd.DataFrame(),
            default_date_inspected_df=None,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
            gaspar_candidate_df=gaspar_candidate_df,
            tri_classification_df=tri_classification_df,
        )

        self.assertEqual(int(summary.loc[0, "flag_jrc"]), 0)
        self.assertEqual(int(summary.loc[0, "flag_gaspar"]), 1)
        self.assertTrue(bool(summary.loc[0, "tri_boundary_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_flood"]), 0)
        self.assertEqual(summary.loc[0, "flag_flood_case"], "none")
        self.assertEqual(summary.loc[0, "flag_flood_source"], "none")
        self.assertEqual(summary.loc[0, "flag_flood_decision_path"], "no_jrc_hit_gaspar_inside_n_tri_not_for")

    def test_build_summary_table_uses_gaspar_riparian_when_outside_n_tri(self) -> None:
        original_points = pd.DataFrame({"point_id": [1], "LAT": [48.8566], "LONG": [2.3522]})
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
                "insee_com": ["75056"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": [pd.NA],
                "start_date": pd.to_datetime([pd.NaT]),
                "end_date": pd.to_datetime([pd.NaT]),
                "raster_path_found": [False],
            }
        )
        gaspar_candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "gaspar_event_uid": ["gaspar-3"],
                "cod_nat_catnat": ["CATNAT-3"],
                "gaspar_start_date": pd.to_datetime(["2020-02-01"]),
                "gaspar_end_date": pd.to_datetime(["2020-02-02"]),
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1],
                "tri_for_hit": [False],
                "tri_boundary_hit": [False],
                "tri_zone_status": ["outside_n_tri"],
                "riparian_hit": [True],
                "flood_risk_area_value": ["medium"],
                "TRI": ["medium"],
                "tri_scenario_codes": ["03Mcc"],
                "tri_scenario_labels": [
                    "Aléa de moyenne probabilité avec prise en compte du changement climatique"
                ],
                "riparian_zone_hit": [False],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=pd.DataFrame(),
            default_date_inspected_df=None,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
            gaspar_candidate_df=gaspar_candidate_df,
            tri_classification_df=tri_classification_df,
        )

        self.assertEqual(int(summary.loc[0, "flag_jrc"]), 0)
        self.assertEqual(int(summary.loc[0, "flag_gaspar"]), 1)
        self.assertTrue(bool(summary.loc[0, "riparian_hit"]))
        self.assertEqual(int(summary.loc[0, "flag_flood"]), 1)
        self.assertEqual(summary.loc[0, "flag_flood_case"], "case_c_gaspar_riparian")
        self.assertEqual(summary.loc[0, "flag_flood_source"], "gaspar")
        self.assertEqual(summary.loc[0, "flag_flood_start_date"], pd.Timestamp("2020-02-01"))
        self.assertEqual(summary.loc[0, "flag_flood_end_date"], pd.Timestamp("2020-02-02"))

    def test_build_summary_table_keeps_gaspar_outside_n_tri_and_riparian_negative(self) -> None:
        original_points = pd.DataFrame({"point_id": [1], "LAT": [48.8566], "LONG": [2.3522]})
        points_with_lau = pd.DataFrame(
            {
                "point_id": [1],
                "excel_row_number": [2],
                "lau_code": ["FR_75056"],
                "lau_code_local": ["75056"],
                "lau_name": ["Paris"],
                "country_code": ["FR"],
                "insee_com": ["75056"],
            }
        )
        candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "event_id": [pd.NA],
                "start_date": pd.to_datetime([pd.NaT]),
                "end_date": pd.to_datetime([pd.NaT]),
                "raster_path_found": [False],
            }
        )
        gaspar_candidate_df = pd.DataFrame(
            {
                "point_id": [1],
                "gaspar_event_uid": ["gaspar-4"],
                "cod_nat_catnat": ["CATNAT-4"],
                "gaspar_start_date": pd.to_datetime(["2020-03-01"]),
                "gaspar_end_date": pd.to_datetime(["2020-03-03"]),
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1],
                "tri_for_hit": [False],
                "tri_boundary_hit": [False],
                "tri_zone_status": ["outside_n_tri"],
                "riparian_hit": [False],
                "flood_risk_area_value": ["low"],
                "TRI": ["low"],
                "tri_scenario_codes": ["04Fai"],
                "tri_scenario_labels": ["Aléa de faible probabilité"],
                "riparian_zone_hit": [False],
            }
        )

        summary = build_summary_table(
            original_points=original_points,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            points_with_lau=points_with_lau,
            candidate_df=candidate_df,
            inspected_df=pd.DataFrame(),
            default_date_inspected_df=None,
            point_buffer_m=40.0,
            surrounding_buffer_km=1.0,
            threshold_cm=0.0,
            study_start=None,
            study_end=None,
            gaspar_candidate_df=gaspar_candidate_df,
            tri_classification_df=tri_classification_df,
        )

        self.assertEqual(int(summary.loc[0, "flag_jrc"]), 0)
        self.assertEqual(int(summary.loc[0, "flag_gaspar"]), 1)
        self.assertEqual(summary.loc[0, "tri_zone_status"], "outside_n_tri")
        self.assertEqual(int(summary.loc[0, "flag_flood"]), 0)
        self.assertEqual(summary.loc[0, "flag_flood_case"], "none")
        self.assertEqual(summary.loc[0, "flag_flood_source"], "none")
        self.assertEqual(
            summary.loc[0, "flag_flood_decision_path"],
            "no_jrc_hit_gaspar_outside_n_tri_and_riparian",
        )

    def test_build_hanze_candidate_sheet_uses_tri_and_riparian_rules(self) -> None:
        hanze_candidate_df = pd.DataFrame(
            {
                "point_id": [1, 2, 3],
                "LAT": [48.1, 48.2, 48.3],
                "LONG": [2.1, 2.2, 2.3],
                "excel_row_number": [2, 3, 4],
                "lau_code": ["FR_00001", "FR_00002", "FR_00003"],
                "lau_name": ["Point A", "Point B", "Point C"],
                "insee_com": ["00001", "00002", "00003"],
                "insee_dep": ["01", "02", "03"],
                "nuts3_code": ["FR101", "FR201", "FR301"],
                "hanze_nuts3_name": ["Dept A", "Dept B", "Dept C"],
                "hanze_event_uid": ["hanze-1", "hanze-2", "hanze-3"],
                "hanze_event_id": ["1", "2", "3"],
                "hanze_start_date": pd.to_datetime(["2020-01-01", "2020-02-01", "2020-03-01"]),
                "hanze_end_date": pd.to_datetime(["2020-01-02", "2020-02-02", "2020-03-02"]),
                "hanze_country_code": ["FR", "FR", "FR"],
                "hanze_country_name": ["France", "France", "France"],
                "hanze_event_type": ["Flood", "Flood", "Flood"],
                "hanze_flood_source": ["River", "River", "River"],
            }
        )
        tri_classification_df = pd.DataFrame(
            {
                "point_id": [1, 2, 3],
                "tri_for_hit": [True, False, False],
                "tri_boundary_hit": [False, True, False],
                "tri_zone_status": ["for", "inside_n_tri_not_for", "outside_n_tri"],
                "riparian_hit": [False, False, True],
            }
        )

        candidate_sheet = build_hanze_candidate_sheet(
            hanze_candidate_df=hanze_candidate_df,
            point_columns=PointColumns(latitude="LAT", longitude="LONG", point_id="point_id", city=None),
            tri_classification_df=tri_classification_df,
        )
        hits_sheet = build_hanze_hits_sheet(candidate_sheet)

        self.assertEqual(candidate_sheet["flood_risk_area_value"].tolist(), ["high", "other", "out"])
        self.assertEqual(candidate_sheet["hanze_spatial_hit"].tolist(), [True, False, True])
        self.assertEqual(
            candidate_sheet["hanze_hit_reason"].tolist(),
            ["tri_for", "not_selected", "riparian_outside_n_tri"],
        )
        self.assertEqual(hits_sheet["point_id"].tolist(), [1, 3])


if __name__ == "__main__":
    unittest.main()
